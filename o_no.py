from openpyxl import load_workbook
from xlrd import open_workbook
from tobool import to_bool

wb = open_workbook('o_no.xls')
ws = wb.sheet_by_index(0)
a = ws.cell(1, 0).value
b = ws.cell(2, 0).value
a_bool = bool(a)
b_bool = bool(b)
print(a, type(a), a_bool)
print(b, type(b), b_bool)

wb = load_workbook('o_no.xlsx', data_only=True)
ws = wb.active
a = ws["A2"].value
b = ws["A3"].value
a_bool = bool(a)
b_bool = bool(b)
print(a, type(a), a_bool)
print(b, type(b), b_bool)
print()

print("Not spreadsheet per se, but what Python does")
print('bool(False)')
print(bool(False))
print('bool("False")')  # no exception!
print(bool("False"))
print('to_bool("False")')
print(to_bool("False"))
print()

b_to_bool = to_bool(b)
print(b_to_bool, type(b_to_bool))
