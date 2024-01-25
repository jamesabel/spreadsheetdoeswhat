from ismain import is_main
from openpyxl import load_workbook
from xlrd import open_workbook


def use_openpyxl(row: int):
    wb = load_workbook('totally_normal_spreadsheet.xlsx', data_only=True)
    ws = wb.active
    print("openpyxl:")
    for column in ('A', 'B', 'C'):
        value = ws[column + str(row+1)].value
        print(value, type(value))


def use_xlrd(row: int):
    wb = open_workbook('totally_normal_spreadsheet.xls')
    ws = wb.sheet_by_index(0)
    print("xlrd:")
    for column in (0, 1, 2):
        value = ws.cell(row, column).value
        print(value, type(value))


def main():
    for row in range(2):
        print(20 * "=")
        use_openpyxl(row)
        print()
        use_xlrd(row)


if is_main():
    main()
